import re
from openpyxl.descriptors.base import Length
import pandas as pd 
import os
from openpyxl import load_workbook

# funcion para convertir de formato YYYY-MM-DD a DD MMMM
def convertir(trin):
    trin.replace("2020-","")
    t= trin.split('-')
    res=''
    if t[0]=='01':
        res=t[1]+" enero"
    elif t[0]=='02':
        res=t[1]+" febrero"
    elif t[0]=='03':
        res=t[1]+" marzo"
    elif t[0]=='04':
        res=t[1]+" abril"
    elif t[0]=='05':
        res=t[1]+" mayo"
    elif t[0]=='06':
        res=t[1]+" junio"
    elif t[0]=='07':
        res=t[1]+" julio"
    elif t[0]=='08':
        res=t[1]+" agosto"
    elif t[0]=='09':
        res=t[1]+" septiembre"
    elif t[0]=='10':
        res=t[1]+" octubre"
    elif t[0]=='11':
        res=t[1]+" noviembre"
    elif t[0]=='12':
        res=t[1]+" diciembre"
    return res

# funcion para extraer la informacion del excel
def extract_info(path):
    # *se usa pandas para extraer el nombre de la primera hoja del excel
    df = pd.ExcelFile(path)
    hoja= df.sheet_names[0]

    print(path)
    # *se usa openpyxl para cargar las tablas de la hoja del excel 
    # !se usa openpyxl ya que pandas no da la facilidad de extraer tablas del excel
    wb = load_workbook(path)
    ws = wb[hoja]

    mapping = {}

    for entry, data_boundary in ws.tables.items():
        #parse la data en el ref boundary
        data = ws[data_boundary]
        #saca la data 
        #el inner list comprehension agarra el valor de cada celda de la tabla
        content = [[cell.value for cell in ent]
                for ent in data
            ]
        header = content[0]
        #el contenido incluiendo los headers
        rest = content[1:]
        #crea dataframe con los nombres de las columnas
        #y adjunta los nombres con la data
        df = pd.DataFrame(rest, columns = header)
        mapping[entry] = df
        # de la tabla se crea una lista en la cual solo se guarda la semana tentativa para finalizar
        # y se devuelve la lista
        dates=[]
        for date in mapping['tabla']['Semana tentativa para finalizar']:
            dates.append(date)
        
        return dates

# funcion para hacer transformacion y limpieza de datos de las fechas
def clean(dates):
    fecha=[]
    for d in dates:
        string= str(d).lower().strip()
# quitar valores que no se pueden traducir
        if string=='none':
            continue
        if string=='27/04/-03/05':
            continue
        if  string.find("(") != -1:
            continue
        if  string.find("sin tentativa") != -1:
            continue
        if  string.find("según acuerdos") != -1:
            continue
        if  string.find(" la unah.") != -1:
            continue
        if  string.find("según calendario académic") != -1:
            continue
        if string.find("quién fine cuándo") != -1:
            continue
        if string.find("considerando") != -1:
            continue
        if string.find("autoridades universitarias") != -1:
            continue
        if string.find("esta ya termina") != -1:
            continue
        if string.find("semana-finalizar el") != -1:
            continue
        if string.find("la recuperacion") != -1:
            continue
        if string.find("dos") != -1:
            continue
        if string.find("d27e") != -1:
            continue
        if "27 abril-1 d27e mayo" in string:
            continue
        if 'durante' in string:
            continue
# quitar los valores que estorban
        if string.find("may ") != -1:
            string= string.replace("may","mayo")

        if string.find("mayo.") != -1:
            string= string.replace("mayo.","mayo")

        if "1a" in string:
            string= string.replace("1a","1")
        
        if string.find("ii") != -1:
            string= string.replace("ii","2")

        if string.find("2.5") != -1:
            string= string.replace("2,5","3")

        if string.find("aprox.") != -1:
            string= string.replace(" aprox.","")

        if string.find("3.5") != -1:
            string= string.replace("3.5","4").strip()

        if string.find("promedio") != -1:
            string= string.replace("promedio","").strip()

        if string.find("del") != -1:
            string= string.replace("del","")

        if string.find("de") != -1:
            string= string.replace(" de "," ")
            string= string.replace("de "," ")
            string= string.replace(" de"," ")

        if string.find(" al ") != -1:
            string= string.replace(" al ","-")

        if string.find(" – ") != -1:
            string= string.replace(" – ","-").strip()

        if string.find("-") != -1:
            string= string.replace(" - ","-").strip()
            string= string.replace("- ","-").strip()
            string= string.replace(" -","-").strip()

        if string.find(" y ") != -1:
            string= string.replace(" y ","-")

        if string.find("2029") != -1:
                    string= string.replace("2029","2020")

        if string.find("2020-") != -1:
            string= convertir(string)   

        if string.find("2020") != -1:
            string= string.replace(" 2020","")
            string= string.replace("/2020","")

        if string.find(" a ") != -1:
            string= string.replace(" a ","-")

        if string.find("  ") != -1:
            string= string.replace("  "," ")
        
        if string.find("primer") != -1:
            string= string.replace("primer","1")
            string= string.replace("primera","1")

        if string.find("segu") != -1:
            string= string.replace("segunda","2")
            string= string.replace("segunada","2")

        if string.find("tercera") != -1:
            string= string.replace("tercera","3")

        if string.find("cuarta") != -1:
            string= string.replace("cuarta","4")

        if string.find("ltima") != -1:
            string= string.replace("última","4")
            string= string.replace("ultima","4")
        
        if string.find("00:00:00") != -1:
            string= string.replace("00:00:00","")
        
        if string.find("sem") != -1:
            string= string.replace("semanas","semana")
            string= string.replace("semana.","semana")
            string= string.replace("samana","semana")
            string= string.replace("samanas","semana")
            string= string.replace("sem","semana")
            string= string.replace("semna","semana")
            string= string.replace("semanaana","semana")
            string= string.replace("semanana","semana")
            string= string.replace("semanaanda","semana")

        if string.find("semana.") != -1:
                    string= string.replace("semana.","").strip()+" mayo"

        if string.find("samanas") != -1:
                    string= string.replace("semanas","semana mayo").strip()

        if string.find("miércoles, ") != -1:
                    string= string.replace("miércoles, ","")
                    string= string.replace(".","")

        if string.find(" o ") != -1:
                    string= string.split('o')[1].strip()

        if string.find("aproximadamente") != -1:
                    string= string.replace("aproximadamente","mayo")

        if string.find("1 semana may") != -1:
                    string= "1 semana mayo"
# cambio de casos especiales
        if string == '0.5':
            string="1 semana mayo"
        if string == '2.5':
            string="3 semana mayo"
        if string == '1.5 semana':
            string="2 semana mayo"
        if string == '2.5 semana':
            string="3 semana mayo"

        if string == "1":
            string="1 semana mayo"
        if string == "2":
            string="2 semana mayo"
        if string == "3":
            string="3 semana mayo"
        if string == "4":
            string="4 semana mayo"

        if string == "1 semana":
            string="1 semana mayo"
        if string == "2 semana":
            string="2 semana mayo"
        if string == "3 semana":
            string="3 semana mayo"
        if string == "4 semana":
            string="4 semana mayo"
        if string == "4 samanas":
            string="4 semana mayo"

        if string == "1a semana mayo":
            string="1 semana mayo"

        if string == "junio 06 2021":
            string="06 junio"

        if 'examenes-reposisciones' in string:
            string= "3 semana mayo"

        if 'i semana' in string:
            string= "1 semana mayo"

        if '2da-3ra' in string:
            string= "3 semana mayo"
        
        if "1 semana may" in string:
            string= string.replace("may","mayo")
            string= "1 semana mayo"

        if '2 semana may' in string:
            string= "2 semana mayo"

        if '2semana' in string:
            string= "2 semana mayo"

        if '4-7/05/20' in string:
            string= "4-7 mayo"

        if '1a' in string:
            string= string.replace('1a','1')
        
        if '30/04/20' in string:
            string= "27-30 abril"

        if 'semana 2 mayo 11,' in string:
            string= "2 semana mayo"

        if '1 1-15 mayo' in string:
            string= "2 semana mayo"

        if '27 abril-' in string:
            string= "4 semana abril"

        if 'semana 29 mayo' in string:
            string= "4 semana mayo"

        if '04 mayo-08 mayo' in string:
            string= "1 semana mayo"

        if '7/04/-03/05' in string:
            string= "1 semana mayo"

        if '3-4 semana' in string:
            string= "4 semana mayo"

        if '30-4-20' in string:
            string= "4 semana abril"

        if '27-30-04-2020' in string:
            string= "4 semana abril"

        if '11 mayo-14 mayo' in string:
            string= "2 semana abril"

        if '18 mayo-21 mayo' in string:
            string= "3 semana abril"

        if string.split(" ")[0]=="semana":
            string= string.split("semana")[0].strip()

        fecha.append(string.strip())
    # for para cambia de "mes #" a "# mes" 
    for x in range(len(fecha)):
        # recorre lista de los valores que no tienen la palabra "semana"
        if 'semana' not in fecha[x]:
            # separa en el espacio y le da vuelta a los valores
            coso= fecha[x].split(" ")
            if coso[0]=='junio':
                fecha[x]= coso[1]+" "+coso[0]
            elif coso[0]=='abril':
                fecha[x]= coso[1]+" "+coso[0]
            elif coso[0]=='mayo':
                fecha[x]= coso[1]+" "+coso[0]
            elif coso[0]=='julio':
                fecha[x]= coso[1]+" "+coso[0]
    res=[]
    # quita todos los valores vacios 
    for x in fecha:
        if x !="":
            res.append(x)
    # funcion para pasar de formato "# mes" o "intervalo fecha mes" a "# semana mes"
    pasarsemana(res)
    return res

# funcion para pasar de formato "# mes" o "intervalo fecha mes" a "# semana mes"
def pasarsemana(lista):
    for x in range(len(lista)):
        # chequea que solo los valores que no estan en formato de # semana mes sean los que se van a cambiar
        if "semana" not in lista[x]:
            num=0
            # separa la fecha en el espacio para poder sabee el numeor de enfrente
            separado= lista[x].split(" ")

            # chequea si el numero es un intervalo o un digito solo
            if "-" in separado[0]:
                # si es un intervalo se agarra el valor mas alto y se convierte en numero 
                sep= separado[0].split("-")
                num= int(sep[1])
            else:
                # si solo es un valor se convierte en numero 
                num= int(separado[0])
            
            # se cambia en posicion de la lista , si el numero es de 
            # 1-7=1 semana; 
            # 8-15=2 semana; 
            # 16-23=3 semana; 
            # 24-31=4 semana
            # y se le concatena el mes que ya tenia 
            if num<8:
                lista[x]="1 semana "+separado[1]
            elif num<16:
                lista[x]="2 semana "+separado[1]
            elif num<24:
                lista[x]="3 semana "+separado[1]
            else:
                lista[x]="4 semana "+separado[1]

# funcion de ayuda par votaciones que devuele la cantidad de votos de la fecha
def votos(lista):
    return lista[1]

def votaciones(lista,opciones):
    cuartaUrna=[]
    # for para recorrer las opciones de fechas que hay y hacer recuento de cuantos votos hay
    # por fecha y guardando en una lista de listas con fromato ["fecha", #votos]
    for op in opciones:
        temp=[op,lista.count(op)]
        cuartaUrna.append(temp)

    # sort cn ayuda de funcion votos para arreglar de mayor a menor los votos por fecha para definir fin del periodo
    cuartaUrna.sort(key=votos,reverse=True)

    # imprimir ganador de votaciones de tentativa de fin del periodo 
    print("Fecha tentativa de fin de periodo: "+ cuartaUrna[0][0] +" con "+ str(cuartaUrna[0][1]) +" votos")
    for x in cuartaUrna:
        print(x)

def main():
    paths='C:/Users/JOHNNIE/Desktop/sistemas coso/info/'
    dates=[]

    # recorre los archivos que se encuentran en el path
    for path , dir, files in os.walk(paths):
        for f in files:
            # se une la direccion del folder con el archivo que se va a leer 
            p= os.path.join(path,f)
            # se extrae del excel la columna de "Semana tentativa para finalizar" y se agrega a una sola lista
            dates+=extract_info(p)
    
    # se trata de normalizar la mayor cantidad de fechas y se agrega a una nueva lista con las fechas ya normalizadas en formato de semanas(1,2,3,4) por mes
    fechas= clean(dates)

    # imprime la cantidad total de fechas y despues la imprime las fechas valores unicos osea las fechas sin repetir valores
    print("total de fechas ", len(fechas))
    print("total de fechas unique", len(set(fechas)))

    # funcion para calcular votos por fechas para definir cuando es el final del periodo
    votaciones(fechas,list(set(fechas)))
    
    



main()